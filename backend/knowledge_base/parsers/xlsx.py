from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from knowledge_base.cleaning import clean_text, clean_title, format_table_row
from knowledge_base.ocr import OCRService
from knowledge_base.parsers.base import BaseDocumentParser
from knowledge_base.schemas import DocumentParseResult, ParsedSection


class XlsxParser(BaseDocumentParser):
    supported_extensions = (".xlsx",)

    def parse(self, file_path: Path, source: str, ocr_service: OCRService | None = None) -> DocumentParseResult:
        workbook = load_workbook(filename=str(file_path), data_only=True)
        sections: list[ParsedSection] = []
        title = clean_title(file_path.stem)

        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            headers: list[str] = []
            header_set = False
            for row_index, row in enumerate(rows, start=1):
                values = list(row)
                if not any(value not in (None, "") for value in values):
                    continue
                if not header_set:
                    headers = [str(value).strip() if value not in (None, "") else "" for value in values]
                    header_set = True
                    continue
                row_text = format_table_row(headers, values)
                if not row_text:
                    continue
                sections.append(
                    ParsedSection(
                        text=f"Sheet: {sheet.title}\nRow {row_index}: {row_text}",
                        title=title,
                        section=sheet.title,
                        extra_metadata={"sheet_name": sheet.title, "row_index": row_index},
                    )
                )

        text = clean_text("\n\n".join(section.text for section in sections))
        return DocumentParseResult(
            file_path=file_path,
            source=source,
            file_name=file_path.name,
            doc_type="xlsx",
            text=text,
            title=title,
            section=title,
            sections=sections,
        )
